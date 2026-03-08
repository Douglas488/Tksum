from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
    NamedStyle,
)
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment


SOURCE_FILE = Path(
    r"C:\Users\Administrator\Desktop\Tksum\Orders_Profits_Tiktok_BRL_20251001-20251031.xlsx"
)
SOURCE_SHEET = ""  # 留空表示使用工作簿中的第一张工作表
OUTPUT_FILE = SOURCE_FILE.with_name("Tk月总结表.xlsx")

SUMMARY_COLUMNS = [
    "店铺",
    "订单金额",
    "平台回款",
    "产品销售金额",
    "销售折扣",
    "佣金",
    "服务费",
    "其他平台费用",
    "运费",
    "退款",
    "调整",
    "产品数量",
    "商品成本",
    "Upseller利润",
    "毛利率",
    "发票 (8%)",
    "耗损 (3%)",
    "样品成本",
    "样品运费",
    "广告充值",
    "人员和其它费用",
    "净利润",
]

SOURCE_UPSELLER_COLUMN = "利润"


def load_source_dataframe(
    source_file: Path, sheet_name: str | int | None
) -> pd.DataFrame:
    if not source_file.exists():
        raise FileNotFoundError(f"源文件不存在：{source_file}")
    with pd.ExcelFile(source_file) as xls:
        available_sheets = xls.sheet_names
        target_sheet = sheet_name
        if isinstance(sheet_name, str) and not sheet_name.strip():
            target_sheet = None
        if target_sheet not in available_sheets:
            target_sheet = available_sheets[0]
            print(
                f"未找到指定工作表，改为使用第一张工作表：{target_sheet}"
            )
        return pd.read_excel(xls, sheet_name=target_sheet)


def ensure_columns(df: pd.DataFrame) -> pd.DataFrame:
    for column in SUMMARY_COLUMNS:
        if column == "店铺":
            continue
        if column not in df.columns and column not in {"Upseller利润", "毛利率", "发票 (8%)", "耗损 (3%)"}:
            df[column] = 0
    return df


def coerce_numeric_columns(df: pd.DataFrame) -> pd.DataFrame:
    numeric_candidates = [
        col for col in SUMMARY_COLUMNS if col != "店铺" and col in df.columns
    ]
    if SOURCE_UPSELLER_COLUMN in df.columns:
        numeric_candidates.append(SOURCE_UPSELLER_COLUMN)
    for column in numeric_candidates:
        df[column] = pd.to_numeric(df[column], errors="coerce")
    return df


def initialize_manual_columns(df: pd.DataFrame) -> pd.DataFrame:
    manual_columns = ["广告充值", "人员和其它费用", "净利润"]
    for column in manual_columns:
        if column in df.columns:
            df[column] = pd.NA
    return df


def build_summary(df: pd.DataFrame) -> pd.DataFrame:
    value_columns = [
        col
        for col in SUMMARY_COLUMNS
        if col not in {"店铺", "Upseller利润", "毛利率", "发票 (8%)", "耗损 (3%)"}
        and col in df.columns
        and col not in {"广告充值", "人员和其它费用", "净利润"}
    ]
    if SOURCE_UPSELLER_COLUMN in df.columns:
        value_columns.append(SOURCE_UPSELLER_COLUMN)

    grouped = (
        df.groupby("店铺", dropna=False)[value_columns]
        .sum(min_count=1)
        .reset_index()
    )

    for manual_col in ("广告充值", "人员和其它费用"):
        if manual_col not in grouped.columns:
            grouped[manual_col] = 0.0

    if SOURCE_UPSELLER_COLUMN in grouped.columns:
        grouped["Upseller利润"] = grouped[SOURCE_UPSELLER_COLUMN]
        grouped = grouped.drop(columns=[SOURCE_UPSELLER_COLUMN])
    else:
        grouped["Upseller利润"] = (
            grouped["平台回款"] - grouped["商品成本"] - grouped["服务费"]
        )
    # 计算毛利率 = Upseller利润 / 订单金额（小数形式，Excel中会显示为百分比）
    grouped["毛利率"] = grouped["Upseller利润"] / grouped["订单金额"].replace(0, np.nan)
    grouped["毛利率"] = grouped["毛利率"].fillna(0)
    grouped["发票 (8%)"] = grouped["订单金额"] * 0.08
    grouped["耗损 (3%)"] = grouped["订单金额"] * 0.03

    if "订单金额" in df.columns:
        zero_order_mask = np.isclose(df["订单金额"].fillna(0), 0)
        if zero_order_mask.any():
            sample_cost = (
                df.loc[zero_order_mask]
                .groupby("店铺", dropna=False)["商品成本"]
                .sum(min_count=1)
            )
            sample_ship = (
                df.loc[zero_order_mask]
                .groupby("店铺", dropna=False)["运费"]
                .sum(min_count=1)
            )
            grouped["样品成本"] = grouped["店铺"].map(sample_cost).fillna(0)
            grouped["样品运费"] = grouped["店铺"].map(sample_ship).fillna(0)

    grouped["净利润"] = (
        grouped["Upseller利润"]
        - grouped["发票 (8%)"]
        - grouped["耗损 (3%)"]
        - grouped["广告充值"]
        - grouped["人员和其它费用"]
    )

    total_row = grouped.drop(columns=["店铺"]).sum(numeric_only=True).to_dict()
    total_row["店铺"] = "合计"
    # 合计行的毛利率需要重新计算：合计Upseller利润 / 合计订单金额（小数形式，Excel中会显示为百分比）
    if total_row["订单金额"] != 0:
        total_row["毛利率"] = total_row["Upseller利润"] / total_row["订单金额"]
    else:
        total_row["毛利率"] = 0
    summary = pd.concat([grouped, pd.DataFrame([total_row])], ignore_index=True)
    
    # 毛利率单独处理，保留4位小数以确保百分比显示精度
    if "毛利率" in summary.columns:
        summary["毛利率"] = summary["毛利率"].fillna(0).round(4)
    
    # 其他数值列保留2位小数
    numeric_cols = summary.select_dtypes(include="number").columns
    if "毛利率" in numeric_cols:
        numeric_cols = numeric_cols.drop("毛利率")
    summary[numeric_cols] = summary[numeric_cols].round(2)

    for manual_col in ("广告充值", "人员和其它费用", "净利润"):
        if manual_col in summary.columns:
            summary[manual_col] = pd.NA

    return summary[SUMMARY_COLUMNS]


def format_excel_worksheet(ws, summary: pd.DataFrame) -> None:
    """美化Excel工作表 - 高端稳重商务风格"""
    # 高端商务风格配色方案 - 深空灰底色，金色字体
    header_fill = PatternFill(start_color="4A4A4A", end_color="4A4A4A", fill_type="solid")  # 深空灰色
    header_font = Font(bold=True, color="FFD700", size=12, name="等线")  # 金色字体，等线字体
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # 数据行交替背景色
    row_fill_even = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")  # 浅灰
    row_fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # 白色
    
    # 合计行样式
    total_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # 中灰色
    total_font = Font(bold=True, size=11, name="等线", color="000000")  # 等线字体
    
    # 数据行字体 - 使用等线字体，现代优雅
    data_font = Font(size=10, name="等线")
    
    # 红色字体（用于毛利率、样品成本、样品运费）- 使用等线字体
    red_font = Font(size=10, name="等线", color="FF0000")  # 红色
    red_font_bold = Font(size=11, name="等线", color="FF0000", bold=True)  # 红色加粗（合计行）
    
    # 精致边框样式
    border_style = Border(
        left=Side(style="thin", color="808080"),
        right=Side(style="thin", color="808080"),
        top=Side(style="thin", color="808080"),
        bottom=Side(style="thin", color="808080"),
    )
    
    # 表头底部加粗边框
    header_bottom_border = Border(
        left=Side(style="thin", color="808080"),
        right=Side(style="thin", color="808080"),
        top=Side(style="thin", color="808080"),
        bottom=Side(style="medium", color="000000"),
    )
    
    center_alignment = Alignment(horizontal="center", vertical="center")
    right_alignment = Alignment(horizontal="right", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")
    
    num_rows = len(summary) + 1  # 包括表头
    num_cols = len(summary.columns)
    
    # 格式化表头
    for col_idx, col_name in enumerate(summary.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = header_bottom_border
        # 设置表头行高
        ws.row_dimensions[1].height = 25
    
    # 格式化数据行
    for row_idx in range(2, num_rows + 1):
        is_total_row = ws.cell(row=row_idx, column=1).value == "合计"
        
        # 数据行交替背景色（合计行除外）
        if not is_total_row:
            row_fill = row_fill_even if row_idx % 2 == 0 else row_fill_odd
        else:
            row_fill = total_fill
        
        for col_idx, col_name in enumerate(summary.columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border_style
            cell.fill = row_fill
            
            # 设置字体样式（特殊列使用红色）
            red_columns = ["毛利率", "样品成本", "样品运费"]
            if col_name in red_columns:
                # 红色列：合计行用红色加粗，普通行用红色
                if is_total_row:
                    cell.font = red_font_bold
                else:
                    cell.font = red_font
            else:
                # 其他列：合计行用普通加粗，普通行用普通字体
                if is_total_row:
                    cell.font = total_font
                else:
                    cell.font = data_font
            
            # 设置对齐方式
            if col_name == "店铺":
                cell.alignment = left_alignment
            elif col_name == "产品数量":
                cell.alignment = center_alignment
            else:
                cell.alignment = right_alignment
            
            # 设置数字格式（跳过空值单元格）
            if cell.value is None or (isinstance(cell.value, str) and cell.value.strip() == ""):
                continue
                
            try:
                if col_name == "毛利率":
                    cell.value = float(cell.value)
                    cell.number_format = "0.00%"
                elif col_name in ["订单金额", "平台回款", "产品销售金额", "Upseller利润", 
                                 "发票 (8%)", "耗损 (3%)", "净利润"]:
                    cell.value = float(cell.value)
                    cell.number_format = "#,##0.00"
                elif col_name in ["销售折扣", "佣金", "服务费", "其他平台费用", "运费", 
                                 "退款", "调整", "商品成本", "样品成本", "样品运费",
                                 "广告充值", "人员和其它费用"]:
                    cell.value = float(cell.value)
                    cell.number_format = "#,##0.00"
            except (ValueError, TypeError):
                pass
        
        # 设置数据行行高（每行只设置一次）
        ws.row_dimensions[row_idx].height = 20
    
    # 设置所有列宽为161像素（约23 Excel单位）
    # Excel列宽单位转换：1 Excel单位 ≈ 7像素，所以161像素 ≈ 23 Excel单位
    column_width_pixels = 92
    column_width_excel = column_width_pixels / 7  # 转换为Excel单位
    
    for col_idx, col_name in enumerate(summary.columns, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = column_width_excel
    
    # 冻结首行
    ws.freeze_panes = "A2"
    
    # 设置打印区域
    ws.print_area = f"A1:{get_column_letter(num_cols)}{num_rows}"
    
    # 添加批注说明（在表格下方）
    note_row_start = num_rows + 2  # 表格下方空一行后开始
    note_font = Font(size=10, name="等线", color="666666", italic=True)
    note_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # 第一条批注
    note_cell_1 = ws.cell(row=note_row_start, column=1)
    note_cell_1.value = "说明1：所有数据以结算日期为准拉取"
    note_cell_1.font = note_font
    note_cell_1.alignment = note_alignment
    ws.merge_cells(f"A{note_row_start}:{get_column_letter(num_cols)}{note_row_start}")
    
    # 第二条批注
    note_cell_2 = ws.cell(row=note_row_start + 1, column=1)
    note_cell_2.value = "说明2：从Upseller导出利润报表的时候样品的所有费用都已经被Upseller计算进去了，此表中样品成本和样品运费不参与净利润计算(样品定义：取订单金额为0的订单为样品订单)"
    note_cell_2.font = note_font
    note_cell_2.alignment = note_alignment
    ws.merge_cells(f"A{note_row_start + 1}:{get_column_letter(num_cols)}{note_row_start + 1}")
    
    # 设置说明行行高
    ws.row_dimensions[note_row_start].height = 20
    ws.row_dimensions[note_row_start + 1].height = 30  # 第二条说明较长，行高稍高


def generate_report(source_file: Path, source_sheet: str | None = None, output_file: Path | None = None) -> Path:
    """生成报表的核心函数"""
    if output_file is None:
        output_file = source_file.with_name("店铺月度总结报表.xlsx")
    
    df = load_source_dataframe(source_file, source_sheet)
    df = ensure_columns(df)
    df = coerce_numeric_columns(df)
    df = initialize_manual_columns(df)
    summary = build_summary(df)
    summary.to_excel(output_file, index=False)
    
    # 美化Excel表格
    wb = load_workbook(output_file)
    ws = wb.active
    
    # 确保毛利率值正确（在美化前先设置）
    if "毛利率" in summary.columns:
        col_idx = list(summary.columns).index("毛利率") + 1
        col_letter = get_column_letter(col_idx)
        for idx, row_num in enumerate(range(2, len(summary) + 2)):
            cell = ws[f"{col_letter}{row_num}"]
            gross_margin_value = summary.iloc[idx]["毛利率"]
            if pd.notna(gross_margin_value):
                cell.value = float(gross_margin_value)
            else:
                cell.value = 0.0
    
    # 应用美化样式
    format_excel_worksheet(ws, summary)
    
    wb.save(output_file)
    wb.close()
    
    return output_file


def main() -> int:
    """命令行模式"""
    output_file = generate_report(SOURCE_FILE, SOURCE_SHEET if SOURCE_SHEET else None, OUTPUT_FILE)
    print(f"已生成美化报表：{output_file}")
    return 0


def run_gui():
    """运行图形用户界面"""
    try:
        from PySide6.QtWidgets import (
            QApplication,
            QMainWindow,
            QWidget,
            QVBoxLayout,
            QHBoxLayout,
            QLabel,
            QLineEdit,
            QPushButton,
            QFileDialog,
            QMessageBox,
            QProgressBar,
            QTextEdit,
            QGroupBox,
        )
        from PySide6.QtCore import Qt, QThread, Signal
        from PySide6.QtGui import QFont, QIcon
    except ImportError:
        print("错误：需要安装 PySide6 才能使用图形界面")
        print("请运行: pip install pyside6")
        return 1
    
    class ReportGeneratorThread(QThread):
        """报表生成线程"""
        finished = Signal(Path)
        error = Signal(str)
        
        def __init__(self, source_file: Path, source_sheet: str | None, output_file: Path):
            super().__init__()
            self.source_file = source_file
            self.source_sheet = source_sheet
            self.output_file = output_file
        
        def run(self):
            try:
                result = generate_report(self.source_file, self.source_sheet, self.output_file)
                self.finished.emit(result)
            except Exception as e:
                self.error.emit(str(e))
    
    class MainWindow(QMainWindow):
        def __init__(self):
            super().__init__()
            self.setWindowTitle("Tk总结生成")
            self.setMinimumSize(700, 500)
            self.source_file = None
            self.output_file = None
            self.worker_thread = None
            self.init_ui()
        
        def init_ui(self):
            central_widget = QWidget()
            self.setCentralWidget(central_widget)
            layout = QVBoxLayout(central_widget)
            layout.setSpacing(15)
            layout.setContentsMargins(20, 20, 20, 20)
            
            # 标题
            title = QLabel("Tk总结生成")
            title_font = QFont()
            title_font.setPointSize(18)
            title_font.setBold(True)
            title.setFont(title_font)
            title.setAlignment(Qt.AlignmentFlag.AlignCenter)
            layout.addWidget(title)
            
            # 文件选择组
            file_group = QGroupBox("文件选择")
            file_layout = QVBoxLayout()
            file_group.setLayout(file_layout)
            
            # 源文件选择
            source_layout = QHBoxLayout()
            source_label = QLabel("源Excel文件:")
            source_label.setMinimumWidth(100)
            self.source_edit = QLineEdit()
            self.source_edit.setPlaceholderText("请选择源Excel文件...")
            self.source_edit.setReadOnly(True)
            source_browse_btn = QPushButton("浏览...")
            source_browse_btn.clicked.connect(self.select_source_file)
            source_layout.addWidget(source_label)
            source_layout.addWidget(self.source_edit)
            source_layout.addWidget(source_browse_btn)
            file_layout.addLayout(source_layout)
            
            # 输出文件选择
            output_layout = QHBoxLayout()
            output_label = QLabel("输出文件:")
            output_label.setMinimumWidth(100)
            self.output_edit = QLineEdit()
            self.output_edit.setPlaceholderText("将保存为: Tk月总结表.xlsx")
            self.output_edit.setReadOnly(True)
            output_browse_btn = QPushButton("浏览...")
            output_browse_btn.clicked.connect(self.select_output_file)
            output_layout.addWidget(output_label)
            output_layout.addWidget(self.output_edit)
            output_layout.addWidget(output_browse_btn)
            file_layout.addLayout(output_layout)
            
            layout.addWidget(file_group)
            
            # 进度条
            self.progress_bar = QProgressBar()
            self.progress_bar.setVisible(False)
            layout.addWidget(self.progress_bar)
            
            # 生成按钮
            self.generate_btn = QPushButton("生成报表")
            self.generate_btn.setMinimumHeight(40)
            generate_font = QFont()
            generate_font.setPointSize(12)
            generate_font.setBold(True)
            self.generate_btn.setFont(generate_font)
            self.generate_btn.clicked.connect(self.generate_report)
            self.generate_btn.setStyleSheet("""
                QPushButton {
                    background-color: #4A4A4A;
                    color: #FFD700;
                    border: none;
                    border-radius: 5px;
                    padding: 10px;
                }
                QPushButton:hover {
                    background-color: #5A5A5A;
                }
                QPushButton:disabled {
                    background-color: #CCCCCC;
                    color: #666666;
                }
            """)
            layout.addWidget(self.generate_btn)
            
            # 状态信息
            self.status_text = QTextEdit()
            self.status_text.setReadOnly(True)
            self.status_text.setMaximumHeight(100)
            self.status_text.setPlaceholderText("状态信息将显示在这里...")
            layout.addWidget(self.status_text)
            
            # 设置窗口样式
            self.setStyleSheet("""
                QMainWindow {
                    background-color: #F5F5F5;
                }
                QGroupBox {
                    font-weight: bold;
                    border: 2px solid #CCCCCC;
                    border-radius: 5px;
                    margin-top: 10px;
                    padding-top: 10px;
                }
                QGroupBox::title {
                    subcontrol-origin: margin;
                    left: 10px;
                    padding: 0 5px;
                }
                QLineEdit {
                    padding: 5px;
                    border: 1px solid #CCCCCC;
                    border-radius: 3px;
                }
                QPushButton {
                    padding: 5px 15px;
                    border: 1px solid #CCCCCC;
                    border-radius: 3px;
                    background-color: white;
                }
                QPushButton:hover {
                    background-color: #E0E0E0;
                }
            """)
        
        def select_source_file(self):
            file_path, _ = QFileDialog.getOpenFileName(
                self,
                "选择源Excel文件",
                str(Path.home()),
                "Excel文件 (*.xlsx *.xlsm *.xls);;所有文件 (*)"
            )
            if file_path:
                self.source_file = Path(file_path)
                self.source_edit.setText(file_path)
                # 自动设置输出文件路径
                if not self.output_file:
                    self.output_file = self.source_file.with_name("Tk月总结表.xlsx")
                    self.output_edit.setText(str(self.output_file))
                self.update_status(f"已选择源文件: {self.source_file.name}")
        
        def select_output_file(self):
            if not self.source_file:
                QMessageBox.warning(self, "警告", "请先选择源文件")
                return
            
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "选择输出文件",
                str(self.source_file.with_name("Tk月总结表.xlsx")),
                "Excel文件 (*.xlsx);;所有文件 (*)"
            )
            if file_path:
                self.output_file = Path(file_path)
                self.output_edit.setText(file_path)
                self.update_status(f"输出文件: {self.output_file.name}")
        
        def update_status(self, message: str):
            self.status_text.append(f"[{pd.Timestamp.now().strftime('%H:%M:%S')}] {message}")
        
        def generate_report(self):
            if not self.source_file or not self.source_file.exists():
                QMessageBox.warning(self, "错误", "请先选择有效的源Excel文件")
                return
            
            if not self.output_file:
                self.output_file = self.source_file.with_name("Tk月总结表.xlsx")
            
            # 禁用按钮，显示进度条
            self.generate_btn.setEnabled(False)
            self.progress_bar.setVisible(True)
            self.progress_bar.setRange(0, 0)  # 不确定进度
            self.update_status("开始生成报表...")
            
            # 创建工作线程
            self.worker_thread = ReportGeneratorThread(
                self.source_file,
                None,  # 使用第一张工作表
                self.output_file
            )
            self.worker_thread.finished.connect(self.on_report_generated)
            self.worker_thread.error.connect(self.on_report_error)
            self.worker_thread.start()
        
        def on_report_generated(self, output_file: Path):
            self.progress_bar.setVisible(False)
            self.generate_btn.setEnabled(True)
            self.update_status(f"✓ 报表生成成功！")
            self.update_status(f"文件位置: {output_file}")
            
            reply = QMessageBox.question(
                self,
                "生成成功",
                f"报表已成功生成！\n\n文件位置:\n{output_file}\n\n是否打开文件？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                import os
                os.startfile(output_file)
        
        def on_report_error(self, error_msg: str):
            self.progress_bar.setVisible(False)
            self.generate_btn.setEnabled(True)
            self.update_status(f"✗ 生成失败: {error_msg}")
            QMessageBox.critical(self, "错误", f"生成报表时发生错误：\n\n{error_msg}")
    
    app = QApplication(sys.argv)
    app.setStyle("Fusion")  # 使用Fusion样式，更现代
    
    window = MainWindow()
    window.show()
    
    return app.exec()


if __name__ == "__main__":
    # 检查是否有命令行参数，如果没有则启动GUI
    if len(sys.argv) > 1 and sys.argv[1] == "--cli":
        sys.exit(main())
    else:
        sys.exit(run_gui())

