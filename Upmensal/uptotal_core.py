# -*- coding: utf-8 -*-
"""
Upseller 月营业额报表 - 无 GUI 核心逻辑，供 Web API 调用。
从指定文件夹读取所有 .xlsx，按店铺汇总（日期、总销售额、有效订单量、有效销售额等），
生成带书签目录的「所有汇总结果.xlsx」。
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import pandas as pd


def _create_bookmark_framework(ws, excel_files):
    """创建书签目录框架"""
    title_cell = ws.cell(row=1, column=1, value="店铺目录 - 点击可快速跳转到对应店铺数据")
    ws.merge_cells("A1:H1")
    title_cell.font = Font(bold=True, size=12, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25

    bookmark_row = 2
    for filename in excel_files:
        file_title = os.path.splitext(filename)[0]
        cell = ws.cell(row=bookmark_row, column=1, value=f"📌 {file_title}")
        ws.merge_cells(f"A{bookmark_row}:H{bookmark_row}")
        cell.font = Font(color="0000FF", underline="single", size=12)
        cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[bookmark_row].height = 20
        bookmark_row += 1

    ws.row_dimensions[bookmark_row].height = 20
    bookmark_row += 1
    for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col].width = 11.88


def _update_bookmark_links(ws, bookmark_positions, bookmark_data_ranges):
    """更新书签链接"""
    bookmark_row = 2
    for filename in bookmark_positions:
        cell = ws.cell(row=bookmark_row, column=1)
        try:
            if filename in bookmark_data_ranges:
                title_row = bookmark_data_ranges[filename]["title_row"]
                cell.hyperlink = f"#{ws.title}!A{title_row}"
            else:
                cell.hyperlink = f"#{ws.title}!A{bookmark_positions[filename]}"
            cell.font = Font(color="0000FF", underline="single")
        except Exception:
            pass
        bookmark_row += 1


def run_merge(folder_path, output_filename="所有汇总结果.xlsx"):
    """
    从 folder_path 读取所有 .xlsx，汇总后保存为 output_filename，返回输出文件路径。
    Excel 需含列：日期、总销售额、有效订单量、有效销售额。
    """
    excel_files = sorted([f for f in os.listdir(folder_path) if f.endswith(".xlsx")])
    if not excel_files:
        raise FileNotFoundError(f"文件夹中未找到 .xlsx 文件：{folder_path}")

    wb = Workbook()
    ws = wb.active
    ws.title = "汇总结果"

    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.3

    _create_bookmark_framework(ws, excel_files)
    data_start_row = len(excel_files) + 4
    current_row = data_start_row
    bookmark_positions = {}
    bookmark_data_ranges = {}
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for filename in excel_files:
        file_path = os.path.join(folder_path, filename)
        try:
            df = pd.read_excel(file_path)
        except Exception as e:
            raise ValueError(f"读取 {filename} 失败: {e}")

        required = ["日期", "总销售额", "有效订单量", "有效销售额"]
        for col in required:
            if col not in df.columns:
                raise ValueError(f"文件 {filename} 缺少列：{col}")

        df["月份"] = pd.to_datetime(df["日期"]).dt.strftime("%Y.%m")
        summary = (
            df.groupby("月份")
            .agg(
                {
                    "总销售额": "sum",
                    "有效订单量": "sum",
                    "有效销售额": "sum",
                }
            )
            .reset_index()
        )
        summary["客单价"] = summary.apply(
            lambda row: round(row["有效销售额"] / row["有效订单量"], 2)
            if row["有效订单量"] > 0
            else 0,
            axis=1,
        )

        file_title = os.path.splitext(filename)[0]

        # 标题行
        ws.cell(row=current_row, column=1, value=file_title)
        ws.merge_cells(f"A{current_row}:H{current_row}")
        cell = ws.cell(row=current_row, column=1)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True, size=14, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")
        cell.border = thin_border
        ws.row_dimensions[current_row].height = 25
        bookmark_positions[file_title] = current_row
        start_row = current_row
        current_row += 1

        # 表头
        headers = ["月份", "总销售额", "有效订单量", "有效销售额", "客单价", "广告费", "广告销售额", "投产比"]
        for col, header in enumerate(headers, 1):
            hc = ws.cell(row=current_row, column=col, value=header)
            hc.font = Font(bold=True, size=12, color="FFFFFF")
            hc.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            hc.alignment = Alignment(horizontal="center", vertical="center")
            hc.border = thin_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # 数据行
        for row_idx, (_, row_data) in enumerate(summary.iterrows()):
            for col, value in enumerate(row_data, 1):
                data_cell = ws.cell(row=current_row, column=col, value=value)
                data_cell.alignment = Alignment(horizontal="center", vertical="center")
                data_cell.border = thin_border
                if row_idx % 2 == 0:
                    data_cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            for col in [6, 7, 8]:
                c = ws.cell(row=current_row, column=col, value=0)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = thin_border
                if row_idx % 2 == 0:
                    c.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            ws.row_dimensions[current_row].height = 20
            current_row += 1

        # 返回目录行
        return_cell = ws.cell(row=current_row, column=1, value="返回店铺目录  ")
        ws.merge_cells(f"A{current_row}:H{current_row}")
        try:
            return_cell.hyperlink = f"#{ws.title}!A1"
            return_cell.font = Font(color="FFFFFF", underline="single", bold=True, size=12)
            return_cell.alignment = Alignment(horizontal="center", vertical="center")
            return_cell.fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
            return_cell.border = thin_border
        except Exception:
            pass
        ws.row_dimensions[current_row].height = 22
        current_row += 1
        current_row += 1

        end_row = current_row - 2
        bookmark_data_ranges[file_title] = {
            "title_row": start_row,
            "start_row": start_row,
            "end_row": end_row,
        }

    _update_bookmark_links(ws, bookmark_positions, bookmark_data_ranges)
    output_path = os.path.join(folder_path, output_filename)
    wb.save(output_path)
    return output_path
