import os
import threading
import pandas as pd
import ttkbootstrap as tb
from ttkbootstrap.constants import *
from tkinter import filedialog, messagebox
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.page import PageMargins
import subprocess
import time
import win32com.client
import sys

class ExcelMergerApp(tb.Window):
    def __init__(self):
        super().__init__(themename="cosmo")  # 可选主题: cosmo, flatly, minty, etc.
        self.title("Excel 批量汇总工具")
        self.geometry("700x620")
        self.resizable(False, False)
        self.folder_path = tb.StringVar()
        self.create_widgets()
        
        # 在右上角添加使用教程按钮
        tutorial_btn = tb.Button(self, text="📖 使用教程", bootstyle=INFO, width=10, command=self.show_tutorial)
        tutorial_btn.place(relx=0.95, rely=0.02, anchor=NE)

    def create_widgets(self):
        # LOGO区域（可自定义图片路径）
        logo_frame = tb.Frame(self, bootstyle=LIGHT)
        logo_frame.pack(fill=X, pady=(18, 0))
        logo_label = tb.Label(logo_frame, text="📊", font=("Segoe UI Emoji", 32))
        logo_label.pack(side=LEFT, padx=(30, 10))
        title_label = tb.Label(logo_frame, text="Excel 批量汇总工具", font=("微软雅黑", 22, "bold"), bootstyle=PRIMARY)
        title_label.pack(side=LEFT, anchor=N)

        # 分割线
        tb.Separator(self, bootstyle=SECONDARY).pack(fill=X, padx=30, pady=(10, 10))

        # 文件夹选择区域
        folder_frame = tb.Frame(self)
        folder_frame.pack(fill=X, padx=30, pady=(0, 10))
        folder_label = tb.Label(folder_frame, text="选择Excel文件夹：", font=("微软雅黑", 12))
        folder_label.pack(side=LEFT)
        folder_entry = tb.Entry(folder_frame, textvariable=self.folder_path, width=38, font=("微软雅黑", 11), state='readonly')
        folder_entry.pack(side=LEFT, padx=8)
        select_btn = tb.Button(folder_frame, text="浏览...", bootstyle=INFO, command=self.select_folder)
        select_btn.pack(side=LEFT)

        # 导出选项区域
        options_frame = tb.Frame(self)
        options_frame.pack(fill=X, padx=30, pady=(10, 8))
        
        # 导出选项
        self.export_html = tb.BooleanVar(value=False)
        self.export_word = tb.BooleanVar(value=False)
        self.export_pdf = tb.BooleanVar(value=False)
        
        export_label = tb.Label(options_frame, text="导出选项：", font=("微软雅黑", 11, "bold"))
        export_label.pack(anchor=W, pady=(0, 5))
        
        html_check = tb.Checkbutton(options_frame, text="导出HTML文件", variable=self.export_html, bootstyle="round-toggle")
        html_check.pack(anchor=W, pady=2)
        
        word_check = tb.Checkbutton(options_frame, text="用Word打开HTML", variable=self.export_word, bootstyle="round-toggle")
        word_check.pack(anchor=W, pady=2)
        
        pdf_check = tb.Checkbutton(options_frame, text="导出PDF文件", variable=self.export_pdf, bootstyle="round-toggle")
        pdf_check.pack(anchor=W, pady=2)

        # 开始按钮
        self.start_btn = tb.Button(self, text="开始汇总并导出", bootstyle=SUCCESS, width=18, command=self.start_merge)
        self.start_btn.pack(pady=(10, 8))

        # 进度条
        self.progress = tb.Progressbar(self, bootstyle=INFO, length=400, mode='determinate')
        self.progress.pack(pady=(0, 8))
        self.progress['value'] = 0

        # 消息卡片
        self.msg_card = tb.LabelFrame(self, text="处理信息", bootstyle=SECONDARY, padding=10)
        self.msg_card.pack(fill=BOTH, expand=True, padx=30, pady=(0, 18))
        self.result_text = tb.Text(self.msg_card, height=8, font=("微软雅黑", 10), wrap='word', state='disabled', relief='flat', background='#f8f9fa')
        self.result_text.pack(fill=BOTH, expand=True)

    def select_folder(self):
        path = filedialog.askdirectory(title="请选择Excel文件所在文件夹", initialdir=".")
        if path:
            self.folder_path.set(path)
            self.clear_result()
            self.progress['value'] = 0

    def start_merge(self):
        folder = self.folder_path.get()
        if not folder:
            messagebox.showwarning("提示", "请先选择Excel文件夹！")
            return
        self.start_btn.config(state='disabled')
        self.progress['value'] = 0
        self.clear_result()
        threading.Thread(target=self.merge_excels, args=(folder,), daemon=True).start()

    def merge_excels(self, folder_path):
        try:
            excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]
            total = len(excel_files)
            if not excel_files:
                self.show_result(f"错误：在文件夹 '{folder_path}' 中没有找到Excel文件\n")
                self.start_btn.config(state='normal')
                return
            
            # 创建新的工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "汇总结果"
            
            # 设置页面属性（A4纸张，居中）
            ws.page_setup.paperSize = 9  # 9 = A4纸张
            ws.page_setup.orientation = 'portrait'  # 纵向
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToHeight = 1
            ws.page_setup.fitToWidth = 1
            
            # 设置页边距（居中）
            ws.page_margins.left = 0.5
            ws.page_margins.right = 0.5
            ws.page_margins.top = 0.5
            ws.page_margins.bottom = 0.5
            ws.page_margins.header = 0.3
            ws.page_margins.footer = 0.3
            
            # 先创建书签目录框架
            bookmark_positions = {}  # 存储书签位置
            bookmark_data_ranges = {}  # 存储每个店铺的数据行数信息
            self.create_bookmark_framework(ws, excel_files)
            
            # 计算数据开始的行号（书签目录后的位置）
            data_start_row = len(excel_files) + 4  # 书签目录标题(1) + 书签链接(len) + 分隔线(1) + 空行(1) + 1
            
            current_row = data_start_row
            
            for idx, filename in enumerate(excel_files, 1):
                file_path = os.path.join(folder_path, filename)
                try:
                    df = pd.read_excel(file_path)
                    df['月份'] = pd.to_datetime(df['日期']).dt.strftime('%Y.%m')
                    summary = df.groupby('月份').agg({
                        '总销售额': 'sum',
                        '有效订单量': 'sum',
                        '有效销售额': 'sum'
                    }).reset_index()
                    
                    # 计算客单价（有效销售额 / 有效订单量）
                    summary['客单价'] = summary.apply(
                        lambda row: round(row['有效销售额'] / row['有效订单量'], 2) if row['有效订单量'] > 0 else 0, 
                        axis=1
                    )
                    
                    file_title = os.path.splitext(filename)[0]
                    
                    # 添加文件名作为标题行（合并A-H列并居中，绿色填充）
                    ws.cell(row=current_row, column=1, value=file_title)
                    ws.merge_cells(f'A{current_row}:H{current_row}')
                    cell = ws.cell(row=current_row, column=1)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(bold=True, size=14, color="FFFFFF", name="华文楷体")
                    # 设置深绿色填充色
                    green_fill = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")
                    cell.fill = green_fill
                    # 设置边框
                    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                       top=Side(style='thin'), bottom=Side(style='thin'))
                    cell.border = thin_border
                    # 设置行高（27像素约等于20点）
                    ws.row_dimensions[current_row].height = 25
                    
                    # 记录书签位置（标题行的绝对位置）
                    bookmark_positions[file_title] = current_row
                    start_row = current_row  # 记录标题行位置
                    current_row += 1
                    
                    # 添加数据标题行（月份、总销售额、有效订单量、有效销售额、客单价、广告费、广告销售额、投产比）
                    headers = ['月份', '总销售额', '有效订单量', '有效销售额', '客单价', '广告费', '广告销售额', '投产比']
                    for col, header in enumerate(headers, 1):
                        header_cell = ws.cell(row=current_row, column=col, value=header)
                        header_cell.font = Font(bold=True, size=12, color="FFFFFF", name="楷体")  # 设置标题行为粗体
                        # 设置浅蓝色背景
                        header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
                        header_cell.fill = header_fill
                        # 设置居中对齐
                        header_cell.alignment = Alignment(horizontal='center', vertical='center')
                        # 设置边框
                        header_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                                   top=Side(style='thin'), bottom=Side(style='thin'))
                    # 设置标题行行高（27像素约等于20点）
                    ws.row_dimensions[current_row].height = 22
                    current_row += 1
                    
                    # 添加汇总数据
                    for row_idx, (_, row_data) in enumerate(summary.iterrows()):
                        for col, value in enumerate(row_data, 1):
                            data_cell = ws.cell(row=current_row, column=col, value=value)
                            # 设置居中对齐
                            data_cell.alignment = Alignment(horizontal='center', vertical='center')
                            # 设置边框
                            data_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                                    top=Side(style='thin'), bottom=Side(style='thin'))
                            # 设置交替行颜色
                            if row_idx % 2 == 0:
                                data_cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
                        
                        # 为新增的列（F、G、H列）添加特殊的数据和边框
                        # F列：广告费（暂时设为0，实际使用时需要从数据源获取）
                        f_cell = ws.cell(row=current_row, column=6, value=0)
                        f_cell.alignment = Alignment(horizontal='center', vertical='center')
                        f_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                            top=Side(style='thin'), bottom=Side(style='thin'))
                        if row_idx % 2 == 0:
                            f_cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
                        
                        # G列：广告销售额（暂时设为0，实际使用时需要从数据源获取）
                        g_cell = ws.cell(row=current_row, column=7, value=0)
                        g_cell.alignment = Alignment(horizontal='center', vertical='center')
                        g_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                            top=Side(style='thin'), bottom=Side(style='thin'))
                        if row_idx % 2 == 0:
                            g_cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
                        
                        # H列：投产比（广告销售额/广告费，暂时设为0）
                        h_cell = ws.cell(row=current_row, column=8, value=0)
                        h_cell.alignment = Alignment(horizontal='center', vertical='center')
                        h_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                            top=Side(style='thin'), bottom=Side(style='thin'))
                        if row_idx % 2 == 0:
                            h_cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
                        
                        # 设置数据行行高（27像素约等于20点）
                        ws.row_dimensions[current_row].height = 20
                        current_row += 1
                    
                    # 添加返回店铺目录按钮（合并A-H列并右对齐）
                    return_cell = ws.cell(row=current_row, column=1, value="🔙 返回店铺目录  ")
                    ws.merge_cells(f'A{current_row}:H{current_row}')
                    try:
                        return_cell.hyperlink = f"#{ws.title}!A1"
                        return_cell.font = Font(color="FFFFFF", underline="single", bold=True, size=12, name="华文楷体")
                        return_cell.alignment = Alignment(horizontal='center', vertical='center')
                        # 设置橙色背景
                        return_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
                        return_cell.fill = return_fill
                        # 设置边框
                        return_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                                  top=Side(style='thin'), bottom=Side(style='thin'))
                        self.show_result(f"已为 '{file_title}' 添加返回首页按钮\n")
                    except Exception as return_error:
                        self.show_result(f"创建返回首页按钮失败 '{file_title}': {return_error}\n")
                    # 设置返回首页按钮行高（27像素约等于20点）
                    ws.row_dimensions[current_row].height = 22
                    current_row += 1
                    
                    # 添加空行分隔
                    current_row += 1
                    
                    # 记录该店铺的数据范围信息
                    end_row = current_row - 1  # 当前行减1（不包括空行）
                    bookmark_data_ranges[file_title] = {
                        'title_row': start_row,  # 店铺标题行
                        'start_row': start_row,
                        'end_row': end_row,
                        'total_rows': end_row - start_row + 1
                    }
                    
                    self.show_result(f"已处理：{filename}\n")
                except Exception as e:
                    self.show_result(f"处理 {filename} 时出错：{e}\n")
                self.progress['value'] = int(idx / total * 100)
            
            # 更新书签链接
            self.update_bookmark_links(ws, bookmark_positions, bookmark_data_ranges, data_start_row)
            
            # 保存文件
            output_file = os.path.join(folder_path, "所有汇总结果.xlsx")
            try:
                wb.save(output_file)
                self.show_result(f"文件已成功保存到：{output_file}\n")
            except Exception as save_error:
                self.show_result(f"保存文件时出错：{save_error}\n")
                # 尝试保存不带超链接的版本
                try:
                    # 移除所有超链接
                    for row in ws.iter_rows():
                        for cell in row:
                            if cell.hyperlink:
                                cell.hyperlink = None
                    
                    backup_file = os.path.join(folder_path, "所有汇总结果_无书签.xlsx")
                    wb.save(backup_file)
                    self.show_result(f"已保存无书签版本到：{backup_file}\n")
                except Exception as backup_error:
                    self.show_result(f"保存备用文件也失败：{backup_error}\n")
            
            self.progress['value'] = 100
            self.show_result("已添加书签跳转功能和返回首页按钮\n")
            self.show_result("导出PDF后可快速定位各店铺数据，并可返回首页\n")
            
            # 开始导出流程
            if self.export_html.get() or self.export_word.get() or self.export_pdf.get():
                self.show_result("开始导出流程...\n")
                self.export_to_html_word_pdf(output_file, folder_path)
            else:
                messagebox.showinfo("完成", f"所有汇总结果已保存到：\n{output_file}\n\n已添加书签跳转功能和返回首页按钮！")
        except Exception as e:
            self.show_result(f"发生错误：{e}\n")
        self.start_btn.config(state='normal')
    
    def create_bookmark_framework(self, ws, excel_files):
        """创建书签目录框架"""
        try:
            # 在文件开头添加书签目录标题
            title_cell = ws.cell(row=1, column=1, value="🏠 店铺目录 - 点击可快速跳转到对应店铺数据")
            # 合并A-H列
            ws.merge_cells(f'A1:H1')
            title_cell.font = Font(bold=True, size=12, color="FFFFFF", name="华文楷体")
            # 设置蓝色背景
            title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            title_cell.fill = title_fill
            # 设置居中对齐
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            # 设置书签目录标题行高（27像素约等于20点）
            ws.row_dimensions[1].height = 25
            
            # 为每个店铺预留书签位置
            bookmark_row = 2
            for filename in excel_files:
                file_title = os.path.splitext(filename)[0]
                cell = ws.cell(row=bookmark_row, column=1, value=f"📌 {file_title}")
                # 合并A-H列
                ws.merge_cells(f'A{bookmark_row}:H{bookmark_row}')
                cell.font = Font(color="0000FF", underline="single", size=12)
                # 设置浅蓝色背景
                link_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                cell.fill = link_fill
                # 设置左对齐
                cell.alignment = Alignment(horizontal='left', vertical='center')
                # 设置书签行高（27像素约等于20点）
                ws.row_dimensions[bookmark_row].height = 20
                bookmark_row += 1
            
            # 设置空行行高（27像素约等于20点）
            ws.row_dimensions[bookmark_row].height = 20
            bookmark_row += 1
            
            # 调整列宽（100像素约等于11.88个字符宽度）
            ws.column_dimensions['A'].width = 11.88
            ws.column_dimensions['B'].width = 11.88
            ws.column_dimensions['C'].width = 11.88
            ws.column_dimensions['D'].width = 11.88
            ws.column_dimensions['E'].width = 11.88
            ws.column_dimensions['F'].width = 11.88
            ws.column_dimensions['G'].width = 11.88
            ws.column_dimensions['H'].width = 11.88
            
            self.show_result("书签目录框架已创建\n")
        except Exception as e:
            self.show_result(f"创建书签框架时出错：{e}\n")
    
    def update_bookmark_links(self, ws, bookmark_positions, bookmark_data_ranges, data_start_row):
        """更新书签链接到正确的位置"""
        try:
            bookmark_row = 2
            for filename, row_position in bookmark_positions.items():
                # 更新书签链接到实际数据位置
                # row_position 现在是绝对位置，直接使用
                cell = ws.cell(row=bookmark_row, column=1)
                try:
                    # 使用实际的数据范围信息
                    if filename in bookmark_data_ranges:
                        data_range = bookmark_data_ranges[filename]
                        title_row = data_range['title_row']  # 店铺标题行
                        
                        # 直接跳转到店铺标题行
                        cell.hyperlink = f"#{ws.title}!A{title_row}"
                        # 设置字体样式
                        cell.font = Font(color="0000FF", underline="single")
                        self.show_result(f"书签 '{filename}' 链接到标题行 A{title_row}\n")
                    else:
                        # 如果没有找到数据范围信息，使用默认方式
                        cell.hyperlink = f"#{ws.title}!A{row_position}"
                        cell.font = Font(color="0000FF", underline="single")
                        self.show_result(f"书签 '{filename}' 链接到 A{row_position} (使用默认方式)\n")
                except Exception as link_error:
                    self.show_result(f"创建书签链接失败 '{filename}': {link_error}\n")
                bookmark_row += 1
            
            self.show_result("书签链接已更新\n")
        except Exception as e:
            self.show_result(f"更新书签链接时出错：{e}\n")

    def show_result(self, msg):
        self.result_text.config(state='normal')
        self.result_text.insert('end', msg)
        self.result_text.see('end')
        self.result_text.config(state='disabled')

    def clear_result(self):
        self.result_text.config(state='normal')
        self.result_text.delete('1.0', 'end')
        self.result_text.config(state='disabled')

    def show_tutorial(self):
        """显示使用教程和项目信息"""
        tutorial_window = tb.Toplevel(self)
        tutorial_window.title("📖 使用说明 - Excel批量汇总工具")
        tutorial_window.geometry("800x600")
        tutorial_window.resizable(True, True)
        
        # 创建滚动文本框
        text_frame = tb.Frame(tutorial_window)
        text_frame.pack(fill=BOTH, expand=True, padx=20, pady=20)
        
        # 创建文本框和滚动条
        text_widget = tb.Text(text_frame, wrap='word', font=("微软雅黑", 10), 
                             background='#f8f9fa', relief='flat', padx=15, pady=15)
        scrollbar = tb.Scrollbar(text_frame, orient='vertical', command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar.set)
        
        text_widget.pack(side=LEFT, fill=BOTH, expand=True)
        scrollbar.pack(side=RIGHT, fill=Y)
        
        # 教程内容
        tutorial_content = """
🏢 项目介绍
═══════════════════════════════════════════════════════════════════════════════

Excel批量汇总工具是用于汇总多个店铺Excel文件的工具。

主要功能：
• 批量处理多个Excel文件，自动汇总各店铺的数据
• 生成带有书签导航的汇总报告，方便快速定位各店铺数据
• 支持导出HTML、Word和PDF格式，满足不同使用需求
• 智能计算客单价等关键业务指标
• 美观的界面设计和用户友好的操作体验

📋 使用前准备
═══════════════════════════════════════════════════════════════════════════════

1. Excel文件格式要求：
   • 文件格式：.xlsx
   • 必须包含以下列名：
     - 日期：订单日期
     - 总销售额：订单总金额
     - 有效订单量：有效订单数量
     - 有效销售额：有效订单金额

2. 文件命名：
   • 建议使用店铺名称作为文件名
   • 例如：xxxx.xlsx

3. 文件夹组织：
   • 将所有需要汇总的Excel文件放在同一个文件夹中
   • 确保文件夹路径不包含特殊字符

🚀 使用步骤
═══════════════════════════════════════════════════════════════════════════════

第一步：选择文件夹
• 点击"浏览..."按钮
• 选择包含Excel文件的文件夹
• 系统会自动识别所有.xlsx文件

第二步：设置导出选项
• 导出HTML文件：生成网页格式的汇总报告
• 用Word打开HTML：自动在Word中打开HTML文件
• 导出PDF文件：将Word文档转换为PDF格式

第三步：开始汇总
• 点击"开始汇总并导出"按钮
• 系统会自动处理所有Excel文件
• 进度条显示处理进度
• 处理信息实时显示在下方文本框中

📊 输出结果
═══════════════════════════════════════════════════════════════════════════════

生成的汇总文件包含：

1. 店铺目录（书签导航）：
   • 位于文件顶部
   • 点击店铺名称可快速跳转到对应数据
   • 每个店铺都有独立的书签链接

2. 各店铺数据：
   • 按月份汇总的营业额数据
   • 包含总销售额、有效订单量、有效销售额、客单价
   • 数据按月份排序显示
   • 交替行颜色，便于阅读

3. 返回导航：
   • 每个店铺数据末尾都有"返回店铺目录"按钮
   • 点击可快速返回目录页面

💡 使用技巧
═══════════════════════════════════════════════════════════════════════════════

• 书签功能：在PDF中点击店铺名称可直接跳转到对应数据
• 返回功能：在PDF中点击"返回店铺目录"可回到首页
• 数据验证：处理前请确保Excel文件格式正确
• 文件备份：建议在处理前备份原始Excel文件
• 批量处理：支持同时处理多个店铺的数据文件

🔧 技术支持
═══════════════════════════════════════════════════════════════════════════════

如果遇到问题：
• 检查Excel文件格式是否符合要求
• 确保文件夹路径正确且包含.xlsx文件
• 查看处理信息中的错误提示
• 确保系统已安装Microsoft Office（用于PDF导出）

版本信息：v1.0
开发语言：Python + ttkbootstrap
支持格式：Excel (.xlsx) → HTML → Word → PDF

═══════════════════════════════════════════════════════════════════════════════
        """
        
        # 插入教程内容
        text_widget.insert('1.0', tutorial_content)
        text_widget.config(state='disabled')  # 设置为只读
        
        # 添加关闭按钮
        close_btn = tb.Button(tutorial_window, text="关闭", bootstyle=SECONDARY, 
                             command=tutorial_window.destroy, width=10)
        close_btn.pack(pady=(0, 20))
        
        # 设置窗口焦点
        tutorial_window.focus_set()
        tutorial_window.grab_set()  # 模态窗口

    def export_to_html_word_pdf(self, excel_file, folder_path):
        """导出Excel到HTML，然后用Word打开并导出PDF"""
        try:
            # 步骤1：导出HTML文件
            if self.export_html.get():
                self.show_result("步骤1：正在导出HTML文件...\n")
                html_file = os.path.join(folder_path, "所有汇总结果.html")
                self.excel_to_html(excel_file, html_file)
                self.show_result(f"HTML文件已保存到：{html_file}\n")
            
            # 步骤2：用Word打开HTML文件
            if self.export_word.get() and self.export_html.get():
                self.show_result("步骤2：正在用Word打开HTML文件...\n")
                word_doc = self.open_html_in_word(html_file)
                self.show_result("HTML文件已在Word中打开\n")
            
            # 步骤3：导出PDF文件
            if self.export_pdf.get() and self.export_word.get():
                self.show_result("步骤3：正在导出PDF文件...\n")
                pdf_file = os.path.join(folder_path, "所有汇总结果.pdf")
                self.word_to_pdf(word_doc, pdf_file)
                self.show_result(f"PDF文件已保存到：{pdf_file}\n")
            
            messagebox.showinfo("导出完成", 
                f"导出流程已完成！\n\n"
                f"Excel文件：{excel_file}\n"
                f"{'HTML文件：' + html_file if self.export_html.get() else ''}\n"
                f"{'PDF文件：' + pdf_file if self.export_pdf.get() else ''}\n\n"
                f"所有文件都保存在：{folder_path}")
                
        except Exception as e:
            self.show_result(f"导出过程中发生错误：{e}\n")
            messagebox.showerror("导出错误", f"导出过程中发生错误：{e}")

    def excel_to_html(self, excel_file, html_file):
        """使用Excel的另存为HTML功能，保持完整格式"""
        try:
            # 使用win32com.client自动化Excel
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            
            # 打开Excel文件
            workbook = excel.Workbooks.Open(os.path.abspath(excel_file))
            
            # 另存为HTML格式（保持所有格式和功能）
            html_path = os.path.abspath(html_file)
            workbook.SaveAs(html_path, FileFormat=44)  # 44 = HTML格式
            
            # 等待文件保存完成
            self.show_result("正在等待Excel保存HTML文件...\n")
            time.sleep(2)  # 等待2秒确保文件完全保存
            
            # 关闭Excel
            try:
                workbook.Close()
                excel.Quit()
            except:
                pass
            
            self.show_result(f"Excel已成功另存为HTML：{html_path}\n")
                
        except Exception as e:
            raise Exception(f"Excel另存为HTML失败：{e}")

    def open_html_in_word(self, html_file):
        """用Word打开HTML文件"""
        try:
            # 使用win32com.client创建Word应用程序实例
            word = win32com.client.Dispatch("Word.Application")
            word.Visible = True
            word.DisplayAlerts = False  # 禁用所有警告对话框
            
            # 等待一下确保Word完全启动
            time.sleep(2)
            
            # 尝试直接打开HTML文件，不指定格式
            try:
                doc = word.Documents.Open(os.path.abspath(html_file))
            except:
                # 如果直接打开失败，尝试指定HTML格式
                doc = word.Documents.Open(
                    FileName=os.path.abspath(html_file),
                    ConfirmConversions=False,
                    ReadOnly=False,
                    AddToRecentFiles=False,
                    Format=8  # HTML格式
                )
            
            # 等待文档完全加载
            time.sleep(2)
            
            # 设置页面属性（A4纸张）
            try:
                doc.PageSetup.PageWidth = 595.28  # A4宽度（点）
                doc.PageSetup.PageHeight = 841.89  # A4高度（点）
                doc.PageSetup.LeftMargin = 72  # 1英寸边距
                doc.PageSetup.RightMargin = 72
                doc.PageSetup.TopMargin = 72
                doc.PageSetup.BottomMargin = 72
            except:
                # 如果页面设置失败，继续执行
                pass
            
            self.show_result("Word文档已打开\n")
            return doc
                
        except Exception as e:
            raise Exception(f"打开HTML文件失败：{e}")

    def word_to_pdf(self, word_doc, pdf_file):
        """将Word文档转换为PDF"""
        try:
            # 导出为PDF
            pdf_path = os.path.abspath(pdf_file)
            
            try:
                word_doc.SaveAs(pdf_path, FileFormat=17)  # 17 = PDF格式
            except:
                # 如果指定格式失败，尝试不指定格式
                word_doc.SaveAs(pdf_path)
            
            # 关闭Word文档和应用程序
            try:
                word_doc.Close()
                word_doc.Application.Quit()
            except:
                pass
            
            self.show_result("PDF文件已成功导出\n")
                
        except Exception as e:
            raise Exception(f"PDF导出失败：{e}")

if __name__ == "__main__":
    app = ExcelMergerApp()
    app.mainloop() 