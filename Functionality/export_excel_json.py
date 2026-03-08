import json
import os
from datetime import datetime
from typing import Any, Dict, List, Optional, Set, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PILImage

# UI deps (prefer ttkbootstrap for现代科技感). 允许在缺失时回退到tkinter。
try:
    import ttkbootstrap as tb  # type: ignore
    from ttkbootstrap.dialogs import Messagebox  # type: ignore
    from ttkbootstrap.widgets import DateEntry  # type: ignore
    import tkinter as tk  # type: ignore
    import tkinter.filedialog as filedialog  # type: ignore
    import tkinter.messagebox as messagebox  # type: ignore
    _USE_TTKB = True
except Exception:
    import tkinter as tk  # type: ignore
    from tkinter import ttk, filedialog, messagebox  # type: ignore
    _USE_TTKB = False


def safe_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_date(value: Any) -> str:
    if value is None or value == "/":
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = safe_str(value)
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except Exception:
            pass
    return text


def sanitize_filename_segment(name: str) -> str:
    if not name:
        return ""
    invalid = '<>:"/\\|?*\n\r\t'
    result = ''.join('_' if ch in invalid else ch for ch in name)
    # Trim spaces and collapse repeated underscores
    result = '_'.join(part for part in result.strip().split('_') if part)
    return result


def export_images(ws, images_dir: str, row_start: int, row_to_sku: Dict[int, str], allow_rows: Optional[Set[int]] = None) -> Dict[int, List[str]]:
    os.makedirs(images_dir, exist_ok=True)
    row_to_files: Dict[int, List[str]] = {}

    # openpyxl stores images in ws._images with anchors
    for idx, img in enumerate(getattr(ws, "_images", []), start=1):
        try:
            anchor = img.anchor
            # Only handle cell anchors
            if hasattr(anchor, "_from"):
                row = anchor._from.row + 1  # zero-based to 1-based
                col = anchor._from.col + 1
            elif hasattr(anchor, "from_"):
                row = anchor.from_.row + 1
                col = anchor.from_.col + 1
            else:
                continue

            if row < row_start:
                continue

            if allow_rows is not None and row not in allow_rows:
                continue

            # Derive extension
            ext = "png"
            if isinstance(img, OpenpyxlImage) and img._data is not None:
                fmt = safe_str(getattr(img, "format", "")).lower()
                if fmt in ("png", "jpeg", "jpg", "bmp", "gif"):
                    ext = "jpg" if fmt == "jpeg" else fmt

            sku_prefix = sanitize_filename_segment(row_to_sku.get(row, ""))
            if sku_prefix:
                filename = f"{sku_prefix}_row{row}_col{col}_{idx}.{ext}"
            else:
                filename = f"row{row}_col{col}_{idx}.{ext}"
            path = os.path.join(images_dir, filename)

            # Save image bytes
            try:
                img_ref = getattr(img, "_data", None)
                if img_ref is not None:
                    with open(path, "wb") as f:
                        f.write(img_ref())
                else:
                    # Try PIL fallback
                    pil_img = getattr(img, "_image", None)
                    if pil_img is not None and isinstance(pil_img, PILImage.Image):
                        pil_img.save(path)
                    else:
                        continue
            except Exception:
                continue

            row_to_files.setdefault(row, []).append(path)
        except Exception:
            continue

    return row_to_files


def read_table(excel_path: str, header_row: int = 1, data_start_row: int = 4) -> Tuple[List[str], List[List[Any]], Dict[int, str]]:
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    headers: Dict[int, str] = {}
    for col in range(1, ws.max_column + 1):
        headers[col] = safe_str(ws.cell(row=header_row, column=col).value)

    header_list = [headers.get(c, f"col_{c}") or f"col_{c}" for c in range(1, ws.max_column + 1)]

    row_to_sku: Dict[int, str] = {}
    for row in range(data_start_row, ws.max_row + 1):
        row_to_sku[row] = safe_str(ws.cell(row=row, column=1).value)

    rows: List[List[Any]] = []
    for row in range(data_start_row, ws.max_row + 1):
        sku_cell = safe_str(ws.cell(row=row, column=1).value)
        row_values = [ws.cell(row=row, column=c).value for c in range(1, ws.max_column + 1)]
        if not sku_cell and not any(safe_str(v) for v in row_values):
            continue
        rows.append(row_values)

    return header_list, rows, row_to_sku


def filter_rows_by_dates(headers: List[str], rows: List[List[Any]], date_col_name: str, date_values: Set[str]) -> List[List[Any]]:
    if date_col_name not in headers:
        return []
    idx = headers.index(date_col_name)
    filtered: List[List[Any]] = []
    for r in rows:
        v = normalize_date(r[idx])
        if v in date_values:
            filtered.append(r)
    return filtered


def export_json_from_rows(headers: List[str], rows: List[List[Any]], images_map: Optional[Dict[int, List[str]]], output_json: str) -> None:
    records: List[Dict[str, Any]] = []
    for row_idx, r in enumerate(rows, start=0):
        record: Dict[str, Any] = {}
        for c_idx, key in enumerate(headers, start=0):
            value = r[c_idx]
            if key == "时间":
                record[key] = normalize_date(value)
            else:
                record[key] = safe_str(value)
        if images_map is not None:
            record["图片文件"] = images_map.get(row_idx, [])
        records.append(record)
    with open(output_json, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)


def export_excel_from_rows(headers: List[str], rows: List[List[Any]], output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    for c_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c_idx, value=h)
    for r_idx, r in enumerate(rows, start=2):
        for c_idx, v in enumerate(r, start=1):
            ws.cell(row=r_idx, column=c_idx, value=v)
    wb.save(output_path)


def run_cli_default():
    excel_path = os.path.join(os.path.dirname(__file__), "新品分配.xlsx")
    images_dir = os.path.join(os.path.dirname(__file__), "images")
    output_json = os.path.join(os.path.dirname(__file__), "output.json")

    headers, rows, row_to_sku = read_table(excel_path)

    # 全量导出图片（按行号）
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    row_start = 4
    allow_rows = set(range(row_start, ws.max_row + 1))
    row_to_images_raw = export_images(ws, images_dir, row_start=row_start, row_to_sku=row_to_sku, allow_rows=allow_rows)

    # 将图片映射到相对行索引
    images_map: Dict[int, List[str]] = {}
    for abs_row, files in row_to_images_raw.items():
        rel_index = abs_row - row_start  # 与 rows 列表对齐
        images_map[rel_index] = [os.path.relpath(p, os.path.dirname(output_json)) for p in files]

    export_json_from_rows(headers, rows, images_map, output_json)
    print(f"导出完成：{output_json}，记录数：{len(rows)}，图片目录：{images_dir}")


def start_gui():
    if _USE_TTKB:
        app = tb.Window(title="Excel 智能筛选导出系统", themename="superhero", position=(50, 50), size=(900, 700))
        container = app
        Message = Messagebox
    else:
        app = tk.Tk()
        app.title("Excel 智能筛选导出系统")
        app.geometry("900x700")
        container = app

    # --- State ---
    state: Dict[str, Any] = {
        "excel_path": "",
        "headers": [],
        "rows": [],
        "row_to_sku": {},
        "date_col": "",
        "unique_dates": [],
        "selected_dates": set(),
    }

    def choose_file():
        path = filedialog.askopenfilename(title="选择Excel文件", filetypes=[("Excel", "*.xlsx;*.xlsm;*.xltx;*.xltm")])
        if not path:
            return
        try:
            headers, rows, row_to_sku = read_table(path)
            state["excel_path"] = path
            state["headers"] = headers
            state["rows"] = rows
            state["row_to_sku"] = row_to_sku

            # 识别可能的日期列（优先匹配"时间"）
            date_candidates = [h for h in headers if "时间" in h or "date" in h.lower()]
            values_preview = []
            if date_candidates:
                state["date_col"] = date_candidates[0]
            else:
                state["date_col"] = headers[0] if headers else ""

            update_headers_dropdown()
            refresh_unique_dates()
            info_var.set(os.path.basename(path))
        except Exception as e:
            if _USE_TTKB:
                Message.show_error("读取失败", str(e))
            else:
                messagebox.showerror("读取失败", str(e))

    def update_headers_dropdown():
        items = state["headers"]
        if _USE_TTKB:
            date_col_combo.configure(values=items)
            if state["date_col"]:
                date_col_combo.set(state["date_col"])
        else:
            date_col_combo["values"] = items
            if state["date_col"]:
                date_col_var.set(state["date_col"])

    def refresh_unique_dates():
        headers = state["headers"]
        rows = state["rows"]
        date_col = state["date_col"]
        if not headers or not rows or not date_col:
            return
        idx = headers.index(date_col)
        uniq: List[str] = []
        seen: Set[str] = set()
        for r in rows:
            d = normalize_date(r[idx])
            if d and d not in seen:
                seen.add(d)
                uniq.append(d)
        uniq.sort()
        state["unique_dates"] = uniq
        state["selected_dates"] = set()
        
        # 清空并重建日期选择区域
        for widget in dates_frame.winfo_children():
            widget.destroy()
        
        # 添加全选/全不选按钮
        select_all_frame = tk.Frame(dates_frame)
        select_all_frame.pack(fill="x", pady=(0, 8))
        
        if _USE_TTKB:
            select_all_btn = tb.Button(select_all_frame, text="全选", bootstyle="outline", command=select_all_dates)
            select_all_btn.pack(side="left", padx=(0, 8))
            deselect_all_btn = tb.Button(select_all_frame, text="全不选", bootstyle="outline", command=deselect_all_dates)
            deselect_all_btn.pack(side="left")
        else:
            select_all_btn = ttk.Button(select_all_frame, text="全选", command=select_all_dates)
            select_all_btn.pack(side="left", padx=(0, 8))
            deselect_all_btn = ttk.Button(select_all_frame, text="全不选", command=deselect_all_dates)
            deselect_all_btn.pack(side="left")
        
        # 创建日期复选框
        for i, date in enumerate(uniq):
            date_frame = tk.Frame(dates_frame)
            date_frame.pack(fill="x", pady=2)
            
            if _USE_TTKB:
                var = tb.BooleanVar()
                cb = tb.Checkbutton(date_frame, text=date, variable=var, command=lambda d=date, v=var: toggle_date(d, v.get()))
                cb.pack(side="left")
            else:
                var = tk.BooleanVar()
                cb = ttk.Checkbutton(date_frame, text=date, variable=var, command=lambda d=date, v=var: toggle_date(d, v.get()))
                cb.pack(side="left")
            
            # 存储变量引用
            if not hasattr(state, 'date_vars'):
                state['date_vars'] = {}
            state['date_vars'][date] = var

    def on_date_col_change(event=None):
        val = date_col_var.get() if not _USE_TTKB else date_col_combo.entry.get()
        state["date_col"] = val
        refresh_unique_dates()

    def toggle_date(date: str, selected: bool):
        if selected:
            state["selected_dates"].add(date)
        else:
            state["selected_dates"].discard(date)
        update_selection_count()

    def select_all_dates():
        for date, var in state.get('date_vars', {}).items():
            var.set(True)
            state["selected_dates"].add(date)
        update_selection_count()

    def deselect_all_dates():
        for date, var in state.get('date_vars', {}).items():
            var.set(False)
        state["selected_dates"].clear()
        update_selection_count()

    def update_selection_count():
        count = len(state["selected_dates"])
        if _USE_TTKB:
            selection_info_var.set(f"已选择 {count} 个日期")
        else:
            selection_info_var.set(f"已选择 {count} 个日期")

    def export_action():
        if not state["excel_path"]:
            (Message.show_warning("提示", "请先选择Excel文件") if _USE_TTKB else messagebox.showwarning("提示", "请先选择Excel文件"))
            return
        chosen = list(state["selected_dates"])
        if not chosen:
            (Message.show_warning("提示", "请选择至少一个日期") if _USE_TTKB else messagebox.showwarning("提示", "请选择至少一个日期"))
            return
        choose_dir = filedialog.askdirectory(title="选择导出目录")
        if not choose_dir:
            return

        headers = state["headers"]
        rows = state["rows"]
        date_col = state["date_col"]
        filtered = filter_rows_by_dates(headers, rows, date_col, set(chosen))

        # 导出Excel
        base = os.path.splitext(os.path.basename(state["excel_path"]))[0]
        out_excel = os.path.join(choose_dir, f"{base}_filtered.xlsx")
        export_excel_from_rows(headers, filtered, out_excel)

        # 导出JSON + 图片（仅所选日期）
        out_json = os.path.join(choose_dir, f"{base}_filtered.json")
        # 计算允许的绝对行集合，以便图片导出
        wb = load_workbook(state["excel_path"], data_only=True)
        ws = wb.active
        header_row = 1
        data_start_row = 4
        # 将过滤行映射回绝对行号
        # 根据 read_table 的行构造方式，rows 与绝对行号一一对应（起始为 data_start_row）
        allow_rows_abs: Set[int] = set()
        headers_list, all_rows, row_to_sku = state["headers"], state["rows"], state["row_to_sku"]
        idx = headers.index(date_col)
        for i, r in enumerate(all_rows, start=0):
            if normalize_date(r[idx]) in set(chosen):
                allow_rows_abs.add(i + data_start_row)

        images_dir = os.path.join(choose_dir, "images")
        row_to_images_raw = export_images(ws, images_dir, row_start=data_start_row, row_to_sku=row_to_sku, allow_rows=allow_rows_abs)
        # 将图片相对映射到 filtered 的相对序号
        images_map: Dict[int, List[str]] = {}
        # 构建从绝对行到在filtered中的相对索引
        abs_to_rel: Dict[int, int] = {}
        rel_counter = 0
        for i, r in enumerate(all_rows, start=0):
            if (i + data_start_row) in allow_rows_abs:
                abs_to_rel[i + data_start_row] = rel_counter
                rel_counter += 1
        for abs_row, files in row_to_images_raw.items():
            rel_idx = abs_to_rel.get(abs_row)
            if rel_idx is None:
                continue
            images_map[rel_idx] = [os.path.relpath(p, choose_dir) for p in files]

        export_json_from_rows(headers, filtered, images_map, out_json)

        msg = f"已导出\nExcel: {out_excel}\nJSON: {out_json}\n共 {len(filtered)} 行"
        (Message.show_info("完成", msg) if _USE_TTKB else messagebox.showinfo("完成", msg))

    # --- UI Layout ---
    pad = 20
    if _USE_TTKB:
        # 主容器
        main_container = tb.Frame(container, padding=pad)
        main_container.pack(fill="both", expand=True)
        
        # 标题区域
        title_frame = tb.Frame(main_container)
        title_frame.pack(fill="x", pady=(0, 20))
        
        title = tb.Label(title_frame, text="🚀 Excel 智能筛选导出系统")
        title.pack(side="left")
        
        subtitle = tb.Label(title_frame, text="专业级数据筛选与导出工具")
        subtitle.pack(side="left", padx=(10, 0))
        
        # 文件选择区域
        file_frame = tb.LabelFrame(main_container, text="📁 文件选择", padding=15, bootstyle="info")
        file_frame.pack(fill="x", pady=(0, 15))
        
        file_btn_frame = tb.Frame(file_frame)
        file_btn_frame.pack(fill="x")
        
        choose_btn = tb.Button(file_btn_frame, text="📂 选择Excel文件", command=choose_file)
        choose_btn.pack(side="left")
        
        info_var = tb.StringVar(value="未选择文件")
        info_lbl = tb.Label(file_btn_frame, textvariable=info_var)
        info_lbl.pack(side="left", padx=(15, 0))
        
        # 列选择区域
        column_frame = tb.LabelFrame(main_container, text="📊 列配置", padding=15, bootstyle="success")
        column_frame.pack(fill="x", pady=(0, 15))
        
        tb.Label(column_frame, text="日期列:").pack(anchor="w")
        date_col_combo = tb.Combobox(column_frame)
        date_col_combo.pack(fill="x", pady=(5, 0))
        date_col_combo.bind("<<ComboboxSelected>>", on_date_col_change)
        
        # 日期选择区域
        dates_frame_container = tb.LabelFrame(main_container, text="📅 日期筛选", padding=15, bootstyle="warning")
        dates_frame_container.pack(fill="both", expand=True, pady=(0, 15))
        
        # 滚动区域
        canvas = tk.Canvas(dates_frame_container, height=200)
        scrollbar = tb.Scrollbar(dates_frame_container, orient="vertical", command=canvas.yview)
        scrollable_frame = tb.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        dates_frame = scrollable_frame
        
        # 选择信息
        selection_info_var = tb.StringVar(value="已选择 0 个日期")
        selection_info = tb.Label(dates_frame_container, textvariable=selection_info_var)
        selection_info.pack(pady=(10, 0))
        
        # 导出区域
        export_frame = tb.Frame(main_container)
        export_frame.pack(fill="x")
        
        export_btn = tb.Button(export_frame, text="🎯 开始导出", command=export_action)
        export_btn.pack(side="left")
        
        # 状态信息
        status_var = tb.StringVar(value="就绪")
        status_lbl = tb.Label(export_frame, textvariable=status_var)
        status_lbl.pack(side="left", padx=(20, 0))
    else:
        # 原生tkinter版本（回退模式）
        frm = ttk.Frame(container, padding=pad)
        frm.pack(fill="both", expand=True)
        
        title = ttk.Label(frm, text="Excel 智能筛选导出系统", font=("Segoe UI", 16, "bold"))
        title.grid(row=0, column=0, columnspan=3, pady=(pad, pad//2), sticky="w")

        choose_btn = ttk.Button(frm, text="选择Excel…", command=choose_file)
        choose_btn.grid(row=1, column=0, sticky="w")
        info_var = tk.StringVar(value="未选择文件")
        info_lbl = ttk.Label(frm, textvariable=info_var)
        info_lbl.grid(row=1, column=1, columnspan=2, sticky="w", padx=(8,0))

        ttk.Label(frm, text="日期列").grid(row=2, column=0, sticky="w", pady=(pad, 4))
        date_col_var = tk.StringVar()
        date_col_combo = ttk.Combobox(frm, textvariable=date_col_var)
        date_col_combo.grid(row=2, column=1, columnspan=2, sticky="we")
        date_col_combo.bind("<<ComboboxSelected>>", on_date_col_change)

        ttk.Label(frm, text="选择日期（可多选）").grid(row=3, column=0, sticky="w", pady=(pad, 4))
        
        # 创建滚动区域
        dates_canvas = tk.Canvas(frm, height=200)
        dates_scrollbar = ttk.Scrollbar(frm, orient="vertical", command=dates_canvas.yview)
        dates_scrollable_frame = ttk.Frame(dates_canvas)
        
        dates_scrollable_frame.bind(
            "<Configure>",
            lambda e: dates_canvas.configure(scrollregion=dates_canvas.bbox("all"))
        )
        
        dates_canvas.create_window((0, 0), window=dates_scrollable_frame, anchor="nw")
        dates_canvas.configure(yscrollcommand=dates_scrollbar.set)
        
        dates_canvas.grid(row=3, column=1, columnspan=2, sticky="nsew")
        dates_scrollbar.grid(row=3, column=3, sticky="ns")
        
        dates_frame = dates_scrollable_frame
        
        # 选择信息
        selection_info_var = tk.StringVar(value="已选择 0 个日期")
        selection_info = ttk.Label(frm, textvariable=selection_info_var)
        selection_info.grid(row=4, column=1, columnspan=2, sticky="w", pady=(5, 0))

        export_btn = ttk.Button(frm, text="导出筛选结果", command=export_action)
        export_btn.grid(row=5, column=0, pady=(pad, 0), sticky="w")

        frm.columnconfigure(2, weight=1)
        frm.rowconfigure(3, weight=1)

    app.mainloop()


if __name__ == "__main__":
    # 默认启动现代GUI；如需命令行全量导出，可运行：python export_excel_json.py --cli
    import sys
    if "--cli" in sys.argv:
        run_cli_default()
    else:
        start_gui()
