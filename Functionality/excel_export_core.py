# -*- coding: utf-8 -*-
"""新品 Nx / 包裹尺寸 Excel 转 JSON（无 GUI），供 Web API 调用。仅表数据，不导出图片。支持按日期筛选。"""
from datetime import datetime
from typing import Any, Dict, List, Optional, Set, Tuple
from openpyxl import load_workbook


def safe_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_date(value: Any) -> str:
    if value is None or value == "/":
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = safe_str(value)
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except Exception:
            pass
    return text


def read_table(excel_path: str, header_row: int = 1, data_start_row: int = 4) -> Tuple[List[str], List[List[Any]], Dict[int, str]]:
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    headers = {}
    for col in range(1, ws.max_column + 1):
        headers[col] = safe_str(ws.cell(row=header_row, column=col).value)
    header_list = [headers.get(c, f"col_{c}") or f"col_{c}" for c in range(1, ws.max_column + 1)]
    row_to_sku = {}
    for row in range(data_start_row, ws.max_row + 1):
        row_to_sku[row] = safe_str(ws.cell(row=row, column=1).value)
    rows = []
    for row in range(data_start_row, ws.max_row + 1):
        sku_cell = safe_str(ws.cell(row=row, column=1).value)
        row_values = [ws.cell(row=row, column=c).value for c in range(1, ws.max_column + 1)]
        if not sku_cell and not any(safe_str(v) for v in row_values):
            continue
        rows.append(row_values)
    return header_list, rows, row_to_sku


def filter_rows_by_dates(
    headers: List[str], rows: List[List[Any]], date_col_name: str, date_values: Set[str]
) -> List[List[Any]]:
    """只保留日期列在 date_values 中的行。"""
    if date_col_name not in headers:
        return []
    idx = headers.index(date_col_name)
    return [r for r in rows if normalize_date(r[idx]) in date_values]


def rows_to_records(headers: List[str], rows: List[List[Any]]) -> List[Dict[str, Any]]:
    """将表头与行数据转为记录列表（不包含图片路径）。"""
    records = []
    for r in rows:
        record = {}
        for c_idx, key in enumerate(headers):
            if c_idx >= len(r):
                record[key] = ""
                continue
            value = r[c_idx]
            if key == "时间":
                record[key] = normalize_date(value)
            else:
                record[key] = safe_str(value)
        records.append(record)
    return records


def get_preview(excel_path: str) -> Optional[Dict[str, Any]]:
    """读取 Excel，返回表头、日期列名、以及该列去重排序后的日期列表，供前端做日期选择。"""
    try:
        header_list, rows, _ = read_table(excel_path)
        if not header_list or not rows:
            return {"headers": header_list, "date_column": "", "dates": []}
        date_col = "时间" if "时间" in header_list else (header_list[0] or "列1")
        if date_col not in header_list:
            return {"headers": header_list, "date_column": date_col, "dates": []}
        idx = header_list.index(date_col)
        seen: Set[str] = set()
        uniq: List[str] = []
        for r in rows:
            if idx < len(r):
                d = normalize_date(r[idx])
                if d and d not in seen:
                    seen.add(d)
                    uniq.append(d)
        uniq.sort()
        return {"headers": header_list, "date_column": date_col, "dates": uniq}
    except Exception:
        return None


def run(
    excel_path: str, selected_dates: Optional[List[str]] = None
) -> Optional[List[Dict[str, Any]]]:
    """读取 Excel，返回记录列表。若提供 selected_dates，只返回日期列在该列表中的行。失败返回 None。"""
    try:
        header_list, rows, _ = read_table(excel_path)
        if not header_list:
            return None
        date_col = "时间" if "时间" in header_list else header_list[0]
        if selected_dates is not None and len(selected_dates) > 0 and date_col in header_list:
            date_set: Set[str] = set(selected_dates)
            rows = filter_rows_by_dates(header_list, rows, date_col, date_set)
        return rows_to_records(header_list, rows)
    except Exception:
        return None
