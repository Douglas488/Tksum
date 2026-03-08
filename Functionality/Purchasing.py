#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel转JSON工具 - 用于转换产品资料表
========================================

功能描述：
    一个功能强大的Excel转JSON转换工具，具有现代化的科技感用户界面，
    专门用于转换产品资料表，支持保留超链接信息。

主要特性：
    🎨 科技感界面：深色主题配合绿色科技风格
    📁 文件选择：支持浏览选择任意Excel文件
    🎯 简化操作：默认生成readable.json，无需选择输出路径
    🔗 超链接保留：完整保留Excel中的超链接信息
    📊 智能输出：自动生成易读格式JSON，包含超链接信息
    ⚡ 实时反馈：转换过程中显示详细进度信息
    🛡️ 错误处理：完善的异常捕获和用户提示

使用步骤：
    1. 启动程序：python excel_to_json.py 或双击 启动工具.bat
    2. 选择文件：点击"浏览"按钮选择Excel文件
    3. 开始转换：点击"🚀 开始转换"按钮
    4. 查看结果：在状态区域查看转换进度和结果

输出文件：
    - readable.json（自动生成）：易读格式的JSON文件，包含超链接信息

Excel表头示例：
    SKU | 品名 | 变量 | 采购链接 | 价格 | 库存
    ----|------|------|----------|------|------
    ABC123 | 产品A | 红色 | https://... | 99.99 | 100

系统要求：
    - Python 3.6+
    - pandas >= 1.5.0
    - openpyxl >= 3.0.0
    - tkinter（通常随Python自带）

故障排除：
    1. tkinter不可用：Linux用户需要安装 python3-tk
    2. 依赖包问题：运行 pip install -r requirements.txt
    3. 文件权限：确保对输出目录有写入权限
    4. Excel文件：确保文件未损坏且未被其他程序占用

版本信息：
    v2.0 - 科技感UI版本
    最后更新：2024年
    开发者：AI Assistant
"""

import pandas as pd
import json
import openpyxl
from openpyxl import load_workbook
import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
from typing import Dict, Any, List

def extract_hyperlinks_from_excel(excel_file_path: str) -> Dict[str, Any]:
    """
    从Excel文件中提取数据和超链接信息
    
    Args:
        excel_file_path: Excel文件路径
        
    Returns:
        包含数据和超链接信息的字典
    """
    try:
        # 使用openpyxl加载工作簿以获取超链接信息
        workbook = load_workbook(excel_file_path, data_only=False)
        
        result = {
            "sheets": {},
            "metadata": {
                "file_name": os.path.basename(excel_file_path),
                "total_sheets": len(workbook.sheetnames)
            }
        }
        
        # 遍历所有工作表
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # 获取工作表数据
            sheet_data = []
            hyperlinks = {}
            
            # 遍历所有单元格
            for row in sheet.iter_rows():
                row_data = []
                for cell in row:
                    cell_value = cell.value
                    
                    # 检查是否有超链接
                    if cell.hyperlink:
                        hyperlink_info = {
                            "target": cell.hyperlink.target,
                            "tooltip": cell.hyperlink.tooltip if cell.hyperlink.tooltip else "",
                            "display_text": str(cell_value) if cell_value else ""
                        }
                        
                        # 将超链接信息存储在单独的字典中
                        cell_address = cell.coordinate
                        hyperlinks[cell_address] = hyperlink_info
                        
                        # 在单元格值中添加超链接标记
                        if cell_value:
                            cell_value = f"{cell_value} [HYPERLINK: {cell.hyperlink.target}]"
                        else:
                            cell_value = f"[HYPERLINK: {cell.hyperlink.target}]"
                    
                    row_data.append(cell_value)
                
                # 只添加非空行
                if any(cell is not None for cell in row_data):
                    sheet_data.append(row_data)
            
            # 存储工作表数据
            result["sheets"][sheet_name] = {
                "data": sheet_data,
                "hyperlinks": hyperlinks,
                "total_rows": len(sheet_data),
                "total_hyperlinks": len(hyperlinks)
            }
        
        return result
        
    except Exception as e:
        print(f"读取Excel文件时出错: {str(e)}")
        return None

def excel_to_json_with_hyperlinks(excel_file_path: str, output_json_path: str = None) -> bool:
    """
    将Excel文件转换为JSON格式，保留超链接信息
    
    Args:
        excel_file_path: 输入Excel文件路径
        output_json_path: 输出JSON文件路径（可选）
        
    Returns:
        转换是否成功
    """
    try:
        # 提取数据
        data = extract_hyperlinks_from_excel(excel_file_path)
        
        if data is None:
            return False
        
        # 确定输出文件路径
        if output_json_path is None:
            base_name = os.path.splitext(excel_file_path)[0]
            output_json_path = f"{base_name}_with_hyperlinks.json"
        
        # 保存为JSON文件
        with open(output_json_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        print(f"✅ 成功将Excel文件转换为JSON: {output_json_path}")
        print(f"📊 包含 {data['metadata']['total_sheets']} 个工作表")
        
        # 显示超链接统计
        total_hyperlinks = 0
        for sheet_name, sheet_info in data["sheets"].items():
            hyperlink_count = sheet_info["total_hyperlinks"]
            total_hyperlinks += hyperlink_count
            if hyperlink_count > 0:
                print(f"🔗 工作表 '{sheet_name}' 包含 {hyperlink_count} 个超链接")
        
        print(f"🔗 总计 {total_hyperlinks} 个超链接")
        
        return True
        
    except Exception as e:
        print(f"❌ 转换过程中出错: {str(e)}")
        return False

def create_readable_json(excel_file_path: str, output_json_path: str = None) -> bool:
    """
    创建更易读的JSON格式，将超链接信息整合到数据中
    
    Args:
        excel_file_path: 输入Excel文件路径
        output_json_path: 输出JSON文件路径（可选）
        
    Returns:
        转换是否成功
    """
    try:
        # 提取数据
        data = extract_hyperlinks_from_excel(excel_file_path)
        
        if data is None:
            return False
        
        # 创建简化的格式，只保留数据
        readable_data = []
        
        for sheet_name, sheet_info in data["sheets"].items():
            sheet_data = sheet_info["data"]
            hyperlinks = sheet_info["hyperlinks"]
            
            # 如果有数据，创建表格式结构
            if sheet_data:
                # 假设第一行是标题行
                headers = sheet_data[0] if sheet_data else []
                rows = sheet_data[1:] if len(sheet_data) > 1 else []
                
                # 转换为字典格式
                for i, row in enumerate(rows):
                    row_dict = {}
                    for j, cell_value in enumerate(row):
                        if j < len(headers):
                            header = headers[j] if headers[j] is not None else f"列{j+1}"
                            
                            # 检查是否有对应的超链接
                            cell_address = f"{chr(65+j)}{i+2}"  # Excel坐标格式
                            if cell_address in hyperlinks:
                                row_dict[header] = {
                                    "value": cell_value,
                                    "hyperlink": hyperlinks[cell_address]
                                }
                            else:
                                row_dict[header] = cell_value
                    
                    readable_data.append(row_dict)
        
        # 确定输出文件路径
        if output_json_path is None:
            output_json_path = "readable.json"
        
        # 保存为JSON文件
        with open(output_json_path, 'w', encoding='utf-8') as f:
            json.dump(readable_data, f, ensure_ascii=False, indent=2)
        
        print(f"✅ 成功创建易读格式JSON: {output_json_path}")
        return True
        
    except Exception as e:
        print(f"❌ 创建易读格式时出错: {str(e)}")
        return False

class ExcelToJsonGUI:
    """Excel转JSON工具 - 科技感UI界面"""
    
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Excel转JSON工具 - 用于转换产品资料表")
        self.root.geometry("800x800")
        self.root.configure(bg='#0a0a0a')
        
        # 文件路径变量
        self.selected_file = tk.StringVar()
        
        # 设置科技感主题
        self.setup_style()
        
        # 创建界面
        self.create_widgets()
        
        # 显示欢迎信息
        self.show_welcome_message()
        
    def setup_style(self):
        """设置现代科幻感样式"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # 现代科幻配色方案
        bg_primary = '#0d1117'      # 深色背景
        bg_secondary = '#161b22'    # 次要背景
        bg_tertiary = '#21262d'     # 第三级背景
        accent_cyan = '#00d4ff'     # 青色强调
        accent_purple = '#7c3aed'   # 紫色强调
        text_primary = '#f0f6fc'    # 主要文字
        text_secondary = '#8b949e'  # 次要文字
        border_color = '#30363d'    # 边框颜色
        
        # 配置现代科幻样式
        style.configure('Modern.TFrame', 
                       background=bg_primary,
                       borderwidth=0)
        
        style.configure('Modern.TLabel', 
                       background=bg_primary, 
                       foreground=text_primary,
                       font=('Segoe UI', 10, 'normal'))
        
        style.configure('Title.TLabel',
                       background=bg_primary,
                       foreground=accent_cyan,
                       font=('Segoe UI', 18, 'bold'))
        
        style.configure('Subtitle.TLabel',
                       background=bg_primary,
                       foreground=text_secondary,
                       font=('Segoe UI', 9, 'normal'))
        
        style.configure('Modern.TButton',
                       background=bg_tertiary,
                       foreground=text_primary,
                       borderwidth=1,
                       focuscolor='none',
                       font=('Segoe UI', 9, 'bold'),
                       relief='flat')
        
        style.configure('Primary.TButton',
                       background=accent_cyan,
                       foreground=bg_primary,
                       borderwidth=0,
                       focuscolor='none',
                       font=('Segoe UI', 10, 'bold'),
                       relief='flat')
        
        style.configure('Secondary.TButton',
                       background=bg_secondary,
                       foreground=text_primary,
                       borderwidth=1,
                       focuscolor='none',
                       font=('Segoe UI', 9, 'normal'),
                       relief='flat')
        
        style.configure('Modern.TEntry',
                       fieldbackground=bg_secondary,
                       foreground=text_primary,
                       borderwidth=1,
                       font=('Segoe UI', 9),
                       relief='flat')
        
        style.configure('Modern.Horizontal.TProgressbar',
                       background=accent_cyan,
                       troughcolor=bg_secondary,
                       borderwidth=0,
                       lightcolor=accent_cyan,
                       darkcolor=accent_cyan)
        
        # 悬停和激活效果
        style.map('Modern.TButton',
                 background=[('active', bg_tertiary),
                           ('pressed', accent_purple)])
        
        style.map('Primary.TButton',
                 background=[('active', '#00b8e6'),
                           ('pressed', '#0099cc')])
        
        style.map('Secondary.TButton',
                 background=[('active', bg_tertiary),
                           ('pressed', accent_purple)])
        
        style.map('Modern.TEntry',
                 fieldbackground=[('focus', bg_tertiary)])
        
    def create_widgets(self):
        """创建现代科幻感界面组件"""
        # 主容器
        main_frame = ttk.Frame(self.root, style='Modern.TFrame')
        main_frame.pack(fill=tk.BOTH, expand=True, padx=30, pady=30)
        
        # 标题区域
        title_frame = ttk.Frame(main_frame, style='Modern.TFrame')
        title_frame.pack(fill=tk.X, pady=(0, 30))
        
        # 主标题
        title_label = ttk.Label(title_frame, 
                               text="Excel转JSON工具", 
                               style='Title.TLabel')
        title_label.pack(anchor=tk.W)
        
        # 副标题
        subtitle_label = ttk.Label(title_frame, 
                                  text="智能转换产品资料表 • 保留超链接信息", 
                                  style='Subtitle.TLabel')
        subtitle_label.pack(anchor=tk.W, pady=(5, 0))
        
        # 示例说明
        example_label = ttk.Label(title_frame, 
                                 text="支持格式：SKU | 品名 | 变量 | 采购链接 | 价格 | 库存", 
                                 style='Subtitle.TLabel')
        example_label.pack(anchor=tk.W, pady=(2, 0))
        
        # 文件选择区域
        file_frame = ttk.Frame(main_frame, style='Modern.TFrame')
        file_frame.pack(fill=tk.X, pady=(0, 25))
        
        # 文件选择标签
        file_label = ttk.Label(file_frame, text="选择Excel文件", style='Modern.TLabel')
        file_label.pack(anchor=tk.W, pady=(0, 8))
        
        # 文件输入区域
        file_input_frame = ttk.Frame(file_frame, style='Modern.TFrame')
        file_input_frame.pack(fill=tk.X)
        
        self.file_entry = ttk.Entry(file_input_frame, 
                                   textvariable=self.selected_file,
                                   style='Modern.TEntry')
        self.file_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 12))
        
        browse_btn = ttk.Button(file_input_frame, 
                               text="浏览文件", 
                               command=self.browse_file,
                               style='Secondary.TButton')
        browse_btn.pack(side=tk.RIGHT)
        
        # 输出说明区域
        output_frame = ttk.Frame(main_frame, style='Modern.TFrame')
        output_frame.pack(fill=tk.X, pady=(0, 25))
        
        # 输出标签
        output_label = ttk.Label(output_frame, text="输出文件", style='Modern.TLabel')
        output_label.pack(anchor=tk.W, pady=(0, 8))
        
        # 输出信息显示
        output_info_frame = ttk.Frame(output_frame, style='Modern.TFrame')
        output_info_frame.pack(fill=tk.X)
        
        self.output_label = ttk.Label(output_info_frame, 
                                     text="readable.json",
                                     style='Subtitle.TLabel')
        self.output_label.pack(anchor=tk.W)
        
        # 输出说明
        output_desc = ttk.Label(output_info_frame, 
                               text="自动生成易读格式JSON，包含超链接信息",
                               style='Subtitle.TLabel')
        output_desc.pack(anchor=tk.W, pady=(2, 0))
        
        # 按钮区域
        button_frame = ttk.Frame(main_frame, style='Modern.TFrame')
        button_frame.pack(pady=(30, 0))
        
        # 主要操作按钮
        convert_btn = ttk.Button(button_frame, 
                                text="开始转换", 
                                command=self.start_conversion,
                                style='Primary.TButton')
        convert_btn.pack(side=tk.LEFT, padx=(0, 15))
        
        # 帮助按钮
        help_btn = ttk.Button(button_frame, 
                             text="使用帮助", 
                             command=self.show_help,
                             style='Secondary.TButton')
        help_btn.pack(side=tk.LEFT)
        
        # 进度条
        self.progress = ttk.Progressbar(main_frame, 
                                       style='Modern.Horizontal.TProgressbar',
                                       mode='indeterminate')
        self.progress.pack(fill=tk.X, pady=(25, 0))
        
        # 状态显示区域
        status_frame = ttk.Frame(main_frame, style='Modern.TFrame')
        status_frame.pack(fill=tk.BOTH, expand=True, pady=(25, 0))
        
        # 状态标签
        status_label = ttk.Label(status_frame, text="转换状态", style='Modern.TLabel')
        status_label.pack(anchor=tk.W, pady=(0, 8))
        
        # 创建文本框显示状态
        self.status_text = tk.Text(status_frame, 
                                  height=8,
                                  bg='#161b22',
                                  fg='#f0f6fc',
                                  font=('Segoe UI', 9),
                                  insertbackground='#00d4ff',
                                  selectbackground='#21262d',
                                  selectforeground='#f0f6fc',
                                  borderwidth=0,
                                  relief='flat',
                                  wrap=tk.WORD)
        self.status_text.pack(fill=tk.BOTH, expand=True)
        
        # 滚动条
        scrollbar = ttk.Scrollbar(status_frame, orient=tk.VERTICAL, command=self.status_text.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.status_text.configure(yscrollcommand=scrollbar.set)
        
    def browse_file(self):
        """浏览选择Excel文件"""
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if file_path:
            self.selected_file.set(file_path)
            # 更新输出文件显示
            self.output_label.config(text="readable.json")
            
            
    def log_message(self, message):
        """在状态区域显示消息"""
        self.status_text.insert(tk.END, f"{message}\n")
        self.status_text.see(tk.END)
        self.root.update_idletasks()
        
    def show_help(self):
        """显示使用帮助"""
        help_text = """
🚀 Excel转JSON工具 - 使用帮助
========================================

📋 使用步骤：
1. 点击"浏览"按钮选择Excel文件
2. 点击"🚀 开始转换"按钮
3. 在状态区域查看转换进度和结果

📁 输出文件：
• readable.json（自动生成）
  - 易读格式的JSON文件
  - 结构化的表格数据
  - 超链接整合到数据中

📊 Excel表头示例：
SKU | 品名 | 变量 | 采购链接 | 价格 | 库存
----|------|------|----------|------|------
ABC123 | 产品A | 红色 | https://... | 99.99 | 100

⚠️ 注意事项：
• 确保Excel文件没有损坏且可正常打开
• 支持 .xlsx 和 .xls 格式
• 程序会自动保留超链接信息
• 输出文件保存在程序运行目录

🔧 故障排除：
• tkinter不可用：Linux用户需要安装 python3-tk
• 依赖包问题：运行 pip install -r requirements.txt
• 文件权限：确保对输出目录有写入权限
• Excel文件：确保文件未被其他程序占用

💡 提示：
• 程序使用多线程处理，转换过程不会阻塞界面
• 实时显示转换进度和详细状态信息
• 完善的错误处理和用户提示
        """
        
        # 创建帮助窗口
        help_window = tk.Toplevel(self.root)
        help_window.title("使用帮助")
        help_window.geometry("650x550")
        help_window.configure(bg='#0d1117')
        
        # 帮助窗口主容器
        help_main_frame = ttk.Frame(help_window, style='Modern.TFrame')
        help_main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # 帮助标题
        help_title = ttk.Label(help_main_frame, 
                              text="使用帮助", 
                              style='Title.TLabel')
        help_title.pack(anchor=tk.W, pady=(0, 15))
        
        # 创建文本框显示帮助信息
        help_text_widget = tk.Text(help_main_frame,
                                  bg='#161b22',
                                  fg='#f0f6fc',
                                  font=('Segoe UI', 9),
                                  insertbackground='#00d4ff',
                                  selectbackground='#21262d',
                                  selectforeground='#f0f6fc',
                                  borderwidth=0,
                                  relief='flat',
                                  wrap=tk.WORD)
        help_text_widget.pack(fill=tk.BOTH, expand=True, pady=(0, 15))
        
        # 插入帮助文本
        help_text_widget.insert(tk.END, help_text)
        help_text_widget.config(state=tk.DISABLED)  # 设为只读
        
        # 添加滚动条
        scrollbar = ttk.Scrollbar(help_main_frame, orient=tk.VERTICAL, command=help_text_widget.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        help_text_widget.configure(yscrollcommand=scrollbar.set)
        
        # 添加关闭按钮
        close_btn = ttk.Button(help_main_frame, 
                              text="关闭", 
                              command=help_window.destroy,
                              style='Primary.TButton')
        close_btn.pack(anchor=tk.E, pady=(10, 0))
        
    def show_welcome_message(self):
        """显示欢迎信息"""
        welcome_text = """欢迎使用Excel转JSON工具！

快速开始：
1. 点击"浏览文件"按钮选择Excel文件
2. 点击"开始转换"按钮
3. 程序会自动生成 readable.json 文件

提示：
• 支持 .xlsx 和 .xls 格式
• 自动保留超链接信息
• 点击"使用帮助"查看详细说明

开始转换您的产品资料表吧！"""
        
        self.log_message(welcome_text)
        
    def start_conversion(self):
        """开始转换过程"""
        if not self.selected_file.get():
            messagebox.showerror("错误", "请先选择Excel文件！")
            return
            
        if not os.path.exists(self.selected_file.get()):
            messagebox.showerror("错误", "选择的文件不存在！")
            return
            
        # 清空状态区域
        self.status_text.delete(1.0, tk.END)
        
        # 开始进度条
        self.progress.start()
        
        # 在新线程中执行转换
        thread = threading.Thread(target=self.convert_file)
        thread.daemon = True
        thread.start()
        
    def convert_file(self):
        """执行文件转换"""
        try:
            excel_file = self.selected_file.get()
            
            self.log_message("🚀 开始处理Excel文件...")
            self.log_message(f"📁 输入文件: {os.path.basename(excel_file)}")
            
            # 提取数据
            self.log_message("📊 正在提取数据和超链接信息...")
            data = extract_hyperlinks_from_excel(excel_file)
            
            if data is None:
                self.log_message("❌ 读取Excel文件失败！")
                self.progress.stop()
                return
    
            self.log_message(f"✅ 成功读取 {data['metadata']['total_sheets']} 个工作表")
            
            # 统计超链接
            total_hyperlinks = 0
            for sheet_name, sheet_info in data["sheets"].items():
                hyperlink_count = sheet_info["total_hyperlinks"]
                total_hyperlinks += hyperlink_count
                if hyperlink_count > 0:
                    self.log_message(f"🔗 工作表 '{sheet_name}' 包含 {hyperlink_count} 个超链接")
            
            self.log_message(f"🔗 总计 {total_hyperlinks} 个超链接")
            
            # 创建易读格式JSON
            self.log_message("📝 正在创建易读格式JSON...")
            success = create_readable_json(excel_file, "readable.json")
            if success:
                self.log_message("✅ 易读格式JSON已保存: readable.json")
            else:
                self.log_message("❌ 创建易读格式失败！")
            
            self.log_message("✨ 转换完成！")
            self.log_message("=" * 50)
            
        except Exception as e:
            self.log_message(f"❌ 转换过程中出错: {str(e)}")
        finally:
            self.progress.stop()
            
    def run(self):
        """运行GUI"""
        self.root.mainloop()

def main():
    """主函数 - 启动GUI界面"""
    app = ExcelToJsonGUI()
    app.run()

if __name__ == "__main__":
    main()
