# -*- coding: utf-8 -*-
"""采购 Excel 转 JSON（无 GUI），供 Web API 调用。输出格式与 Purchasing.py 的 readable 一致。"""
import os
from typing import Dict, Any, List
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def extract_hyperlinks_from_excel(excel_file_path: str) -> Dict[str, Any]:
    """从 Excel 提取数据与超链接，返回可序列化的字典。"""
    try:
        workbook = load_workbook(excel_file_path, data_only=False)
        result = {
            "sheets": {},
            "metadata": {
                "file_name": os.path.basename(excel_file_path),
                "total_sheets": len(workbook.sheetnames),
            },
        }
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_data = []
            hyperlinks = {}
            for row in sheet.iter_rows():
                row_data = []
                for cell in row:
                    cell_value = cell.value
                    if cell.hyperlink:
                        hyperlink_info = {
                            "target": cell.hyperlink.target,
                            "tooltip": getattr(cell.hyperlink, "tooltip", "") or "",
                            "display_text": str(cell_value) if cell_value else "",
                        }
                        hyperlinks[cell.coordinate] = hyperlink_info
                        cell_value = f"{cell_value} [HYPERLINK: {cell.hyperlink.target}]" if cell_value else f"[HYPERLINK: {cell.hyperlink.target}]"
                    row_data.append(cell_value)
                if any(c is not None for c in row_data):
                    sheet_data.append(row_data)
            result["sheets"][sheet_name] = {
                "data": sheet_data,
                "hyperlinks": hyperlinks,
                "total_rows": len(sheet_data),
                "total_hyperlinks": len(hyperlinks),
            }
        return result
    except Exception:
        return None


def to_readable_list(data: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    将 extract_hyperlinks_from_excel 的返回转为与 Purchasing.py readable 一致的格式：
    每行一个对象，键为表头，有超链接的单元格为 { "value", "hyperlink": { "target", "tooltip", "display_text" } }。
    """
    readable_data: List[Dict[str, Any]] = []
    for sheet_name, sheet_info in data.get("sheets", {}).items():
        sheet_data = sheet_info.get("data", [])
        hyperlinks = sheet_info.get("hyperlinks", {})
        if not sheet_data:
            continue
        headers = sheet_data[0]
        rows = sheet_data[1:]
        for i, row in enumerate(rows):
            row_dict: Dict[str, Any] = {}
            for j, cell_value in enumerate(row):
                if j >= len(headers):
                    break
                header = headers[j] if headers[j] is not None else f"列{j+1}"
                # Excel 行号：第 0 行是表头，数据从第 2 行开始
                excel_row = i + 2
                col_letter = get_column_letter(j + 1)
                cell_address = f"{col_letter}{excel_row}"
                if cell_address in hyperlinks:
                    row_dict[header] = {
                        "value": cell_value,
                        "hyperlink": hyperlinks[cell_address],
                    }
                else:
                    row_dict[header] = cell_value
            readable_data.append(row_dict)
    return readable_data
